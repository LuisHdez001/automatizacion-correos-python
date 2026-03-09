import os
import subprocess
import re
import shutil
from openpyxl import load_workbook
from datetime import datetime
from openpyxl import load_workbook

excel = "Curso Basico 2026.xlsx"

wb = load_workbook(excel)

# por ahora usamos el mes actual
mes_actual = datetime.now().strftime("%B")

meses = {
    "January": "Enero",
    "February": "Febrero",
    "March": "Marzo",
    "April": "Abril",
    "May": "Mayo",
    "June": "Junio",
    "July": "Julio",
    "August": "Agosto",
    "September": "Septiembre",
    "October": "Octubre",
    "November": "Noviembre",
    "December": "Diciembre"
}

hoja = wb[meses[mes_actual]]

carpeta = "correos/basico"
for archivo in os.listdir(carpeta):
    excel = "Curso Basico 2026.xlsx"

    wb = load_workbook(excel)
    hoja = wb["Febrero"]


for archivo in os.listdir(carpeta):

    # 1 Ignorar lo que NO sea txt
    if not archivo.endswith(".txt"):
        continue

    ruta_completa = os.path.join(carpeta, archivo)

    print("Leyendo:", archivo)

    with open(ruta_completa, "r", encoding="utf-8") as f:
        contenido = f.read()

    # 2 Ignorar correos vacíos
    if not contenido.strip():
        print("Correo vacío, se omite.")
        shutil.move(
            ruta_completa,
            os.path.join("correos/procesados", archivo)
        )
        continue

    # 3 DESPUÉS ya puedes trabajar con el contenido
    buscar_correo = re.search(r'[\w\.-]+@[\w\.-]+', contenido)

    if buscar_correo:
        email = buscar_correo.group(0)
    else:
        email = ""
        
        

    print("Contenido del correo:")
    print(contenido)
    

    resultado = subprocess.run(
        ["python", "extraer.py"],
        input=contenido,
        text=True,
        capture_output=True
    )

    print("Resultado IA:")
    print(resultado.stdout)

    resultado_ia = resultado.stdout.strip()
    lineas = resultado_ia.split("\n")

    for datos in lineas:
        if "|" not in datos:
            continue

        partes = datos.split("|")

        if len(partes) < 3:
            print("Formato incorrecto, se omite.")
            continue

        nombre = partes[0].strip()
        empresa = partes[1].strip()
        puesto = "|".join(partes[2:]).strip()

        hoja.append([
            "",
            "",
            puesto,
            nombre,
            empresa,
            email
        ])

    # guardar cambios
    wb.save(excel)

    # mover archivo a procesados
    destino = os.path.join("correos/procesados", archivo)
    shutil.move(ruta_completa, destino)

    print("Movido a procesados:", archivo)
    print("----------------------")

